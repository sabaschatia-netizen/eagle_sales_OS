"""
Eagle — capa de datos.

Lee el export "cruce" (5 hojas: PRODUCTIVITY, CHECKOUT, ADS, CHURN, MD) y
calcula el TABLERO DE TRIAGE por palanca (Ads, Markdown, Churn) -- no un
funnel de conversión, sino una clasificación mutuamente excluyente de
toda la cartera elegible en un momento dado, con el reloj corriendo
contra la fecha real de HOY (no contra la ventana del Excel), porque el
archivo se actualiza a diario y el sistema mide en vivo.

Formato esperado del Excel, 5 hojas leídas POR NOMBRE (case-insensitive):
  PRODUCTIVITY, CHECKOUT, ADS, CHURN, MD

Todas las funciones reciben los DataFrames ya cargados -- no leen el
disco directamente, así la UI decide si viene de upload o de un archivo
local.
"""

import os
import re

import numpy as np
import pandas as pd
import streamlit as st


# ─────────────────────────────────────────────────────────────
# CARGA
# ─────────────────────────────────────────────────────────────

AM_UNICO = "(único)"

# ── LOGIN ── App de un solo usuario, credenciales fijas (pedido explícito
# de Sabas) -- a diferencia de Wingman, que valida contra una lista de
# Farmers activos y deriva la clave del correo, acá no hace falta nada
# de eso: un solo email, una sola clave, las dos fijas.
LOGIN_EMAIL = "sabas.ramirez@rappi.com"
LOGIN_PASSWORD = "eagle.sabas"


def check_password(email, password):
    """Identificación con correo + contraseña fijos -- no es
    autenticación fuerte (sin hash, sin límite de intentos), es una
    barrera de "por seguridad de datos" ante quien no tenga el link
    directo, mismo espíritu que ya usa Wingman con su propio login."""
    return str(email).strip().lower() == LOGIN_EMAIL and str(password) == LOGIN_PASSWORD

SHEET_ALIASES = {
    "productivity": "productivity",
    "checkout": "checkout",
    "ads": "ads",
    "churn": "churn",
    "md": "md",
}

OUTREACH_COLUMNS = [
    "Brand", "Primer WhatsApp", "Primera Llamada", "Segundo WhatsApp",
    "Segunda Llamada", "Correo Personalizado", "Tercer Llamada",
    "Tercer WhatsApp", "Cuarta Llamada",
]
OUTREACH_ESTADOS = ["No necesario", "No contactado", "Rechazado", "Entró a Pipeline", "Cerrado"]


@st.cache_data(ttl="10m", show_spinner=False)
def load_outreach(file_like_or_path):
    """
    Carga las 3 hojas de seguimiento de Outreach (Ads/MD/Churn) -- las
    genera esta misma app como plantilla, pero se editan a mano fuera de
    acá (Sabas les carga la tipificación en Excel/Sheets), así que hay
    que leerlas tal cual vuelvan.

    BÚSQUEDA DEFENSIVA para la hoja de Churn: "OUTREACH ADS" y
    "OUTREACH MD" se encuentran por nombre sin drama, pero la de Churn
    NO tiene ningún texto identificable en su nombre real -- en la
    primera entrega de este archivo, la hoja de Churn se llamaba
    literalmente "Hoja 10" (el nombre default que le puso Excel/Sheets,
    nunca renombrada a mano). Por eso esa hoja se identifica por
    ESTRUCTURA (mismas 9 columnas exactas de Outreach) entre las que
    sobran después de ya haber apartado Ads y MD por nombre -- no
    confiar en su nombre, porque no está garantizado.

    Hoja faltante = DataFrame vacío con las columnas correctas, nunca
    revienta.
    """
    xls = pd.ExcelFile(file_like_or_path)
    encontradas = {"ads": None, "md": None, "churn": None}
    usadas = set()

    for name in xls.sheet_names:
        upper = name.strip().upper()
        if "OUTREACH" in upper and "CHURN" in upper:
            encontradas["churn"] = name
            usadas.add(name)
        elif "OUTREACH" in upper and ("ADS" in upper or "NEVER" in upper):
            encontradas["ads"] = name
            usadas.add(name)
        elif "OUTREACH" in upper and ("MARKDOWN" in upper or upper.endswith(" MD") or upper == "MD"):
            encontradas["md"] = name
            usadas.add(name)

    if encontradas["churn"] is None:
        for name in xls.sheet_names:
            if name in usadas:
                continue
            try:
                probe = pd.read_excel(xls, sheet_name=name, nrows=0)
                if list(probe.columns) == OUTREACH_COLUMNS:
                    encontradas["churn"] = name
                    usadas.add(name)
                    break
            except (ValueError, KeyError):
                continue

    out = {}
    for key, sheet_name in encontradas.items():
        df = pd.read_excel(xls, sheet_name=sheet_name) if sheet_name else pd.DataFrame(columns=OUTREACH_COLUMNS)
        # Cada hoja de Outreach trae una leyenda pegada abajo de los
        # datos reales (mismo archivo que arma esta app -- filas vacías
        # + "Leyenda" + 3 líneas de texto explicativo). Sin filtrarla,
        # esas líneas llegan a la UI como si fueran marcas fantasma. Se
        # descartan filas sin ningún valor válido de OUTREACH_ESTADOS en
        # las columnas B-I -- una marca real siempre tiene "No
        # necesario" como mínimo en todas, la leyenda nunca tiene nada.
        if not df.empty:
            cols_estado = [c for c in OUTREACH_COLUMNS[1:] if c in df.columns]
            valido = df[cols_estado].isin(OUTREACH_ESTADOS).any(axis=1)
            df = df[df["Brand"].notna() & valido].reset_index(drop=True)
        out[key] = df
    return out


RECUPERADAS_COLUMNS = ["Brand", "Propuesta reformulada", "Status"]


@st.cache_data(ttl="10m", show_spinner=False)
def load_recuperadas(file_like_or_path):
    """
    Carga las 2 hojas de Recuperaciones (Ads/MD) -- mismo patrón que
    load_outreach: las genera esta app como plantilla, se editan a mano
    afuera, se leen tal cual vuelvan. Búsqueda por nombre ("RECUPERADAS"
    + "ADS"/"MD") -- esta vez el archivo real sí trae nombres
    identificables ("RECUPERADAS ads", "RECUPERADAS MD"), a diferencia de
    la hoja de Churn de Outreach que no tenía nombre -- igual se deja un
    fallback estructural por si en algún momento se rompe la convención
    de nombre, mismo criterio defensivo que ya se aplicó ahí.

    Filtra la leyenda pegada al final (mismo mecanismo que
    load_outreach). Hoja faltante = DataFrame vacío, nunca revienta.
    """
    xls = pd.ExcelFile(file_like_or_path)
    encontradas = {"ads": None, "md": None}
    usadas = set()

    for name in xls.sheet_names:
        upper = name.strip().upper()
        if "RECUPERAD" not in upper:
            continue
        if "ADS" in upper:
            encontradas["ads"] = name
            usadas.add(name)
        elif "MD" in upper or "MARKDOWN" in upper:
            encontradas["md"] = name
            usadas.add(name)

    for key in ("ads", "md"):
        if encontradas[key] is not None:
            continue
        for name in xls.sheet_names:
            if name in usadas:
                continue
            try:
                probe = pd.read_excel(xls, sheet_name=name, nrows=0)
                if list(probe.columns) == RECUPERADAS_COLUMNS:
                    encontradas[key] = name
                    usadas.add(name)
                    break
            except (ValueError, KeyError):
                continue

    out = {}
    for key, sheet_name in encontradas.items():
        df = pd.read_excel(xls, sheet_name=sheet_name) if sheet_name else pd.DataFrame(columns=RECUPERADAS_COLUMNS)
        if not df.empty:
            # Fila real = tiene Brand Y (Propuesta reformulada o Status
            # con algo) -- la leyenda solo tiene texto en la columna A.
            valido = df["Propuesta reformulada"].notna() | df["Status"].notna()
            df = df[df["Brand"].notna() & valido].reset_index(drop=True)
            # tabla_lateral() (la misma que ya usa Leads) espera "Brand
            # ID" y "Brand Name" separados, no un solo texto "Brand" --
            # se separa acá para reusarla tal cual, sin duplicar el
            # renderizado de tabla para esta sección nueva.
            df["Brand ID"] = df["Brand"].apply(_brand_key)
            df["Brand Name"] = df["Brand"]
        out[key] = df
    return out


def funnel_recuperadas_niveles(df, gmv_map=None):
    """
    3 niveles de Recuperaciones -- pedido explícito de Sabas, con 2
    lecturas de su spec que quedan documentadas acá por si hay que
    corregirlas:

      DECISIÓN 1: "barra interna: Propuesta reformulada y Sin gestionar"
      del bloque 1 se lee como: Reformulada (Propuesta reformulada="Si")
      vs. Sin Gestionar (="No" o vacío) -- el bloque en sí YA es
      "Rechazadas" (universo completo de la hoja), la barra no repite
      ese nombre como segmento.

      DECISIÓN 2: la spec nombra "Rechazo definitivo" como destino de la
      NO-recuperación, pero el Excel real solo tiene 3 valores posibles
      en Status: Pipeline / Cerrado / Perdida. Se lee "Perdida" =
      "Rechazo definitivo" (mismo concepto, otro nombre) -- no hay un
      cuarto valor en los datos que pueda ser distinto a estos 2.

    Nivel 1 -- Rechazadas = Reformulada + Sin Gestionar (universo entero
      de la hoja).
    Nivel 2 -- Reformulada = Recuperadas (Status en Pipeline/Cerrado) +
      Rechazo definitivo (Status = Perdida).
    Nivel 3 -- Recuperadas = Pipeline + Cerrada.

    gmv_map (nuevo): la hoja de Recuperadas no trae GMV -- se pega desde
    dl.gmv_lookup(), la misma fuente que ya usa Leads, cruzando por
    Brand ID. Marca sin match = $0 (no revienta, no inventa un número).
    """
    gmv_map = gmv_map or {}
    df = df.copy()
    df["GMV"] = df["Brand ID"].apply(lambda k: gmv_map.get(k, 0.0))

    reformulada = df["Propuesta reformulada"].astype(str).str.strip().str.lower() == "si"
    sin_gestionar = ~reformulada

    recuperada = df["Status"].isin(["Pipeline", "Cerrado"])
    rechazo_def = df["Status"] == "Perdida"

    n_reformulada = int(reformulada.sum())
    n_sin_gestionar = int(sin_gestionar.sum())
    n_recuperadas = int((reformulada & recuperada).sum())
    n_rechazo_def = int((reformulada & rechazo_def).sum())
    n_pipeline = int((reformulada & (df["Status"] == "Pipeline")).sum())
    n_cerrada = int((reformulada & (df["Status"] == "Cerrado")).sum())

    def segs(pares):
        total = sum(n for _, n in pares) or 1
        return [{"label": lab, "n": n, "pct": n / total * 100} for lab, n in pares]

    def status_legible(row):
        if not reformulada[row.name]:
            return "Sin Gestionar"
        if row["Status"] == "Perdida":
            return "Rechazo definitivo"
        if row["Status"] in ("Pipeline", "Cerrado"):
            return row["Status"]
        return "Reformulada"  # Si=Si pero Status todavía vacío (recién marcado)

    df_status = df.copy()
    df_status["Status_legible"] = df_status.apply(status_legible, axis=1)

    rechazadas_tbl = df_status.assign(Status=df_status["Status_legible"].where(
        df_status["Status_legible"] == "Sin Gestionar", "Reformulada"))
    reformulada_tbl = df_status[reformulada].assign(
        Status=lambda d: d["Status_legible"].where(d["Status_legible"] == "Rechazo definitivo", "Recuperada"))
    recuperadas_tbl = df_status[reformulada & recuperada].assign(Status=df_status["Status_legible"])

    return [
        {
            "key": "rechazadas", "titulo": "Rechazadas", "total": len(df),
            "sub": "100%",
            "segmentos": segs([("Reformulada", n_reformulada), ("Sin Gestionar", n_sin_gestionar)]),
            "tabla": rechazadas_tbl,
        },
        {
            "key": "reformulada", "titulo": "Reformulada", "total": n_reformulada,
            "sub": f"{(n_reformulada / len(df) * 100) if len(df) else 0:.1f}% de Rechazadas",
            "segmentos": segs([("Recuperada", n_recuperadas), ("Rechazo definitivo", n_rechazo_def)]),
            "tabla": reformulada_tbl,
        },
        {
            "key": "recuperadas", "titulo": "Recuperadas", "total": n_recuperadas,
            "sub": f"{(n_recuperadas / len(df) * 100) if len(df) else 0:.1f}% de Rechazadas",
            "segmentos": segs([("Pipeline", n_pipeline), ("Cerrada", n_cerrada)]),
            "tabla": recuperadas_tbl,
        },
    ]


@st.cache_data(ttl="10m", show_spinner=False)
def _detectar_formato_pct(file_like_or_path, hoja, col_nombre):
    """
    Lee el number_format REAL de cada celda de una columna con openpyxl
    (pandas.read_excel no lo trae) -- devuelve una lista de bool, una por
    fila de datos, True si la celda tiene formato de porcentaje ('%' en
    el patrón de formato, ej. '0%', '0.00%').

    Si algo falla (hoja no existe, columna no encontrada, archivo-like
    ya consumido por una lectura anterior), devuelve una lista de puros
    False -- el resto del pipeline sigue funcionando exactamente como
    antes de este fix, solo sin el detalle nuevo.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_like_or_path, data_only=True)
        if hoja not in wb.sheetnames:
            hoja_real = next((s for s in wb.sheetnames if s.strip().upper() == hoja.upper()), None)
            if not hoja_real:
                return []
            hoja = hoja_real
        ws = wb[hoja]
        headers = [c.value for c in ws[1]]
        if col_nombre not in headers:
            return []
        col_idx = headers.index(col_nombre) + 1
        return [
            "%" in (ws.cell(row=r, column=col_idx).number_format or "")
            for r in range(2, ws.max_row + 1)
        ]
    except Exception:
        return []


def load_cruce(file_like_or_path):
    """
    Carga el Excel de 5 hojas por NOMBRE (no por ancho -- el formato viejo
    de 2 hojas quedó reemplazado). Case-insensitive y con espacios
    recortados, por si el nombre real trae mayúsculas/espacios distintos.
    Devuelve dict: {"productivity", "checkout", "ads", "churn", "md"} --
    hoja faltante = DataFrame vacío, no revienta.
    """
    xls = pd.ExcelFile(file_like_or_path)
    lower_map = {name.strip().lower(): name for name in xls.sheet_names}

    hojas = {}
    for key, alias in SHEET_ALIASES.items():
        real_name = lower_map.get(alias)
        hojas[key] = pd.read_excel(xls, sheet_name=real_name) if real_name else pd.DataFrame()

    productivity = hojas["productivity"]
    checkout = hojas["checkout"]

    if "Date" in productivity.columns:
        productivity["Date"] = pd.to_datetime(productivity["Date"], errors="coerce")
    if "Fecha" in checkout.columns:
        checkout["Fecha"] = pd.to_datetime(checkout["Fecha"], errors="coerce")
    if "Fecha Reactivación" in productivity.columns:
        productivity["Fecha Reactivación"] = pd.to_datetime(
            productivity["Fecha Reactivación"], errors="coerce"
        )
    if "Brand ID" in checkout.columns:
        checkout["brand_key"] = checkout["Brand ID"].apply(_brand_key)
    # BUG REAL ENCONTRADO (pedido explícito de Sabas: "el cierre me está
    # mostrando $0"). pandas.read_excel() SOLO trae el valor numérico
    # crudo de cada celda -- nunca su formato visual. La columna
    # Presupuesto de Checkout mezcla dos tipos de dato en la misma
    # columna sin ningún texto que los distinga:
    #   - Montos en pesos: celda con formato "[$$]#,##0" -> 50000.0
    #   - Porcentajes: celda con formato "0%" -> 0.1 (osea 10%, pero el
    #     valor crudo que devuelve pandas es 0.1, IDÉNTICO a como se
    #     vería 0.1 peso -- no hay forma de diferenciarlos mirando solo
    #     el número). _parse_presupuesto_valor() buscaba el símbolo "%"
    #     en el texto, pero ese símbolo nunca llega -- vive en el
    #     FORMATO de la celda de Excel, no en el valor. Por eso $0.1
    #     pesos ÷ 1450 (tasa USD) redondeaba a $0, silenciosamente,
    #     para CADA fila que en realidad era un %.
    # FIX: se abre el archivo una segunda vez con openpyxl (que sí puede
    # leer number_format) SOLO para la columna Presupuesto, y se guarda
    # el resultado en una columna nueva "_presupuesto_es_pct" -- el
    # resto del pipeline pandas sigue igual, sin tener que reabrir el
    # archivo en cada función que toque Presupuesto.
    if "Presupuesto" in checkout.columns:
        formatos_pct = _detectar_formato_pct(file_like_or_path, "CHECKOUT", "Presupuesto")
        if len(formatos_pct) == len(checkout):
            checkout["_presupuesto_es_pct"] = formatos_pct
        else:
            # Longitud no calza (raro, pero mejor no asignar mal que
            # desalinear filas) -- se sigue igual que antes de este fix.
            checkout["_presupuesto_es_pct"] = False
    # BUG REAL ENCONTRADO (agosto 2026, octavo ajuste): "Tipo de
    # Contratacion" en Checkout trae "Adquisicion " con un espacio en
    # blanco al final -- una comparación exacta contra "Adquisicion"
    # (sin espacio) no matcheaba NINGUNA fila, aunque value_counts()
    # mostraba 96 filas visualmente idénticas. Se recorta acá, en la
    # carga, para que ningún otro lugar del código tenga que acordarse
    # de este detalle.
    if "Tipo de Contratacion" in checkout.columns:
        checkout["Tipo de Contratacion"] = checkout["Tipo de Contratacion"].astype(str).str.strip()
    if "FARMER" in checkout.columns:
        checkout["FARMER"] = checkout["FARMER"].astype(str).str.strip()
    if "Code" in productivity.columns:
        productivity["brand_key"] = productivity["Code"].apply(_brand_key)

    # App de un solo Account Manager: si el export no trae columna de
    # Farmer (o viene vacía), se rellena con un valor único en vez de
    # reventar -- antes esto tiraba "No encontré la columna Farmer".
    if "Farmer" not in productivity.columns or productivity["Farmer"].dropna().empty:
        productivity["Farmer"] = AM_UNICO
    if not checkout.empty and ("FARMER" not in checkout.columns or checkout["FARMER"].dropna().empty):
        checkout["FARMER"] = AM_UNICO

    hojas["productivity"] = productivity
    hojas["checkout"] = checkout
    return hojas


def _asegurar_brand_key(prod):
    """
    Devuelve `prod` con columna `brand_key` garantizada.

    BUG REAL ENCONTRADO (agosto 2026): `funnel_ads`, `funnel_md` y
    `funnel_churn` recalculaban `brand_key` leyendo `prod["Code"]`
    directo, sin la misma protección que `load_cruce` sí aplica (crear
    la columna solo `if "Code" in productivity.columns`). Si el Excel
    real no trae esa columna con ese nombre exacto, esto tiraba
    `KeyError: 'Code'` -- como pasó en el deploy. Ahora se reutiliza el
    `brand_key` que `load_cruce` ya calculó de forma segura y, si por
    algún motivo no llegó (hoja vacía, columna faltante), se cae a
    vacío en vez de reventar la app.
    """
    if "brand_key" in prod.columns:
        return prod
    if "Code" in prod.columns:
        prod["brand_key"] = prod["Code"].apply(_brand_key)
    else:
        prod["brand_key"] = ""
    return prod


def _brand_key(value):
    """Normaliza 'AR104267', '104267', 104267.0, '104308 - Granados Bar Ar'
    -> '104267'/'104308' (solo el número, sin importar prefijo de país,
    sufijo de nombre, ni si viene como texto o numérico) -- las 5 hojas
    no usan el mismo formato de ID para la misma marca."""
    if pd.isna(value):
        return ""
    m = re.search(r"(\d+)", str(value))
    return m.group(1) if m else ""


def _parse_pct(value):
    """'21,97 %' -> 21.97 ; '0,0 %' -> 0.0 ; NaN/vacío -> NaN.
    Las hojas ADS y MD traen los % como texto con coma decimal, no como
    número -- hay que parsearlos antes de poder comparar contra 0."""
    if pd.isna(value):
        return float("nan")
    s = str(value).strip().replace("%", "").replace(",", ".").strip()
    if not s:
        return float("nan")
    try:
        return float(s)
    except ValueError:
        return float("nan")


def _parse_money(value):
    """'$3.324' -> 3324.0 (el punto acá es separador de miles, no decimal
    -- distinto criterio que _parse_pct, que sí usa coma decimal).
    NaN/vacío -> 0.0 (para que sume/ordene bien sin casos especiales)."""
    if pd.isna(value):
        return 0.0
    s = str(value).strip().replace("$", "").replace(".", "").replace(",", ".").strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


@st.cache_data(show_spinner=False)
def gmv_lookup(md_df):
    """Diccionario {brand_key: gmv} construido desde MD.'GMV TOTAL $' --
    se usa como referencia de GMV para las 3 tablas de triage (Ads, MD,
    Churn), no solo para la de MD, porque es la única hoja del cruce que
    trae GMV por marca. Una marca que no aparezca en MD (ej. ya cerrada
    hace tiempo y fuera del export de MD) simplemente no tiene GMV de
    referencia -- se le asigna 0, no revienta."""
    d = md_df.copy()
    if d.empty or "BRAND ID" not in d.columns or "GMV TOTAL $" not in d.columns:
        return {}
    d["brand_key"] = d["BRAND ID"].apply(_brand_key)
    d["gmv"] = d["GMV TOTAL $"].apply(_parse_money)
    return dict(zip(d["brand_key"], d["gmv"]))


def universo_mensual_path(palanca, carpeta="data", hoy=None):
    """Ruta del archivo que acumula el universo fijo del mes en curso
    para esta palanca -- el nombre incluye año-mes, así un mes nuevo
    arranca con archivo nuevo (universo vacío) automáticamente, sin
    necesitar lógica de "reset" aparte."""
    hoy = pd.Timestamp(hoy) if hoy is not None else pd.Timestamp.now()
    return os.path.join(carpeta, f"universo_{palanca}_{hoy.strftime('%Y-%m')}.csv")


def leer_universo_mensual(path):
    """Lee el universo acumulado tal cual está en disco ahora mismo, sin
    modificarlo -- para mostrarlo en la UI (sidebar) y que la
    acumulación deje de ser invisible. A propósito SIN @st.cache_data:
    tiene que reflejar el archivo real al instante, incluso justo
    después de borrarlo con "Reiniciar universo del mes"."""
    if not os.path.exists(path):
        return {}
    try:
        prev = pd.read_csv(path, dtype=str)
        return dict(zip(prev["brand_key"], prev["nombre"]))
    except (pd.errors.EmptyDataError, KeyError):
        return {}


def actualizar_universo_mensual(nuevos_keys_nombres, path):
    """
    Acumula el universo de "Prospectados" contra lo que ya estaba
    guardado en `path` -- pedido explícito de Sabas: "la totalidad de
    prospectados debe ser fija todo el mes". Una marca que calificó un
    día (0% Att. Bookings / Markdown vacío) sigue contando el resto del
    mes aunque al día siguiente ya no cumpla el filtro crudo -- si se
    recalculara desde cero cada vez, una marca que empieza a atribuir
    aunque sea un poco desaparecería del tablero a mitad de camino, sin
    haber llegado a Cerrado ni a Rechazado.

    `nuevos_keys_nombres`: dict {brand_key: nombre} de los que califican
    HOY según el filtro crudo de la hoja. Se agregan los que falten al
    archivo acumulado -- nunca se sacan los que ya estaban. Devuelve el
    dict acumulado completo (lo que hay que usar como universo real).
    """
    if os.path.exists(path):
        try:
            prev = pd.read_csv(path, dtype=str)
            acumulado = dict(zip(prev["brand_key"], prev["nombre"]))
        except (pd.errors.EmptyDataError, KeyError):
            acumulado = {}
    else:
        acumulado = {}

    hubo_nuevos = False
    for k, n in nuevos_keys_nombres.items():
        if k and k not in acumulado:
            acumulado[k] = n
            hubo_nuevos = True

    if hubo_nuevos or not os.path.exists(path):
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        pd.DataFrame({"brand_key": list(acumulado.keys()), "nombre": list(acumulado.values())}).to_csv(
            path, index=False
        )

    return acumulado


def resolver_am(productivity_df):
    """Account Manager de la app. Es de un solo usuario, así que se
    resuelve solo (el valor más frecuente de la columna Farmer) -- no hay
    selector, pedido explícito de Sabas: "solo soy yo"."""
    if "Farmer" not in productivity_df.columns:
        return AM_UNICO
    vals = productivity_df["Farmer"].dropna()
    return vals.mode().iloc[0] if not vals.empty else AM_UNICO


def farmers_disponibles(productivity_df):
    if "Farmer" not in productivity_df.columns:
        return []
    return sorted(productivity_df["Farmer"].dropna().unique().tolist())


def farmer_display(email):
    """'sabas.ramirez@rappi.com' -> 'Sabas Ramirez' -- mismo criterio que
    usa Wingman para mostrar el nombre en la session pill."""
    if email == AM_UNICO:
        return "Account Manager"
    if not email or "@" not in str(email):
        return str(email)
    local = str(email).split("@")[0]
    return " ".join(p.capitalize() for p in local.split("."))


def farmer_initials(email):
    """'sabas.ramirez@rappi.com' -> 'SR' -- para el avatar de la session pill."""
    name = farmer_display(email)
    partes = name.split()
    if not partes:
        return "?"
    if len(partes) == 1:
        return partes[0][:2].upper()
    return (partes[0][0] + partes[1][0]).upper()


# ═════════════════════════════════════════════════════════════
# VENTANA Y DÍAS HÁBILES
# ═════════════════════════════════════════════════════════════
# Pedido explícito de Sabas (décimo ajuste): la ventana se mide SIEMPRE
# desde el primer día hábil del mes en curso hasta hoy, y la temperatura
# de los leads (Caliente/Frío) también se cuenta en días HÁBILES, no
# calendario -- un lead contactado el viernes no debe "enfriarse" por el
# fin de semana.

def primer_dia_habil_mes(hoy=None):
    hoy = pd.Timestamp(hoy) if hoy is not None else pd.Timestamp.now().normalize()
    d = hoy.replace(day=1)
    while d.weekday() >= 5:  # 5=sáb, 6=dom
        d += pd.Timedelta(days=1)
    return d.normalize()


def dias_habiles_entre(desde, hasta):
    """Días hábiles transcurridos entre dos fechas (excluye el día de
    inicio, cuenta el de llegada). Devuelve None si alguna es NaT."""
    if pd.isna(desde) or pd.isna(hasta):
        return None
    d0 = pd.Timestamp(desde).normalize()
    d1 = pd.Timestamp(hasta).normalize()
    if d1 < d0:
        return 0
    return int(np.busday_count(d0.date(), d1.date()))


# ═════════════════════════════════════════════════════════════
# FUNNEL DE 4 NIVELES (Ads / Markdown)
# ═════════════════════════════════════════════════════════════
# Estructura y reglas pedidas explícitamente por Sabas (décimo ajuste,
# con referencia visual incluida). Cada nivel es un subconjunto real del
# anterior y la suma de los segmentos de su barra interna es EXACTAMENTE
# el total de ese nivel ("regla de oro").
#
#   Nivel 1 -- Base prospectada = Contactado + No Contactado + Sin Gestionar
#       Contactado    : ¿Contactado?=SI Y la columna de la palanca (Ads /
#                       Markdown) también =SI -- se habló de ESTA palanca.
#       No Contactado : ¿Contactado?=NO (se intentó, no se logró).
#       Sin Gestionar : ¿Contactado?=SI pero SIN tocar la palanca, o la
#                       marca no tiene fila / el campo viene vacío.
#
#   Nivel 2 -- Contactado = Pipeline + Rechazado
#       Rechazado : tiene "No activo" Y más de 5 días hábiles desde el
#                   primer contacto real con la palanca.
#       Pipeline  : todo el resto de los contactados.
#
#   Nivel 3 -- Pipeline = Caliente + Frío  (días HÁBILES)
#       Caliente : hasta 3 días hábiles
#       Frío     : más de 3 días hábiles
#
#   Nivel 4 -- Cierre : los que aparecen en CHECKOUT (terminal).
#
# DECISIÓN DE DISEÑO, marcada explícita porque la spec no la cubría: un
# contactado con "No activo" pero de 5 días hábiles o menos NO cae en
# Rechazado (la regla exige >5 días) -- queda en Pipeline, clasificado
# por su antigüedad como cualquier otro. Y un contactado de más de 5 días
# SIN "No activo" tampoco cae en Rechazado -- queda en Frío. Cualquier
# otra lectura dejaría marcas sin bucket y rompería la regla de oro.

FUNNEL_ORDEN_L1 = ["Contactado", "No Contactado", "Sin Gestionar"]
FUNNEL_ORDEN_L2 = ["Pipeline", "Rechazado"]
FUNNEL_ORDEN_L3 = ["Caliente", "Frío"]

COLS_TRIAGE = ["Brand ID", "Brand Name", "GMV", "Status"]

UMBRAL_CALIENTE_HABILES = 3
UMBRAL_RECHAZO_HABILES = 5

# 1 USD = 1450 ARS -- pedido explícito de Sabas para convertir el GMV
# (que la tabla siempre muestra en USD) a ARS, que es la moneda en la
# que vienen los rangos de la hoja RECOMMENDED BUDGETS.
TASA_USD_ARS = 1450


def _canal_desde_medio(medio):
    """Traduce el 'Medio de Contacto' crudo de Productivity a uno de los
    3 canales que pide la tabla de Contactados. En la fuente real solo
    aparecen dos valores -- 'Amazon Connect' (el link cae en
    connect.aws, o sea llamada telefónica) y 'Treble' (link cae en
    sales.treble.ai, la plataforma de WhatsApp Business que usa el
    equipo) -- no hay un tercer valor para correo. Se asume Gmail como
    canal por defecto para cualquier contacto sin uno de esos dos medios
    (registro sin 'Medio de Contacto' o con un valor distinto)."""
    if medio == "Amazon Connect":
        return "Llamada"
    if medio == "Treble":
        return "WhatsApp"
    return "Gmail"


def _clasificar_marca(key, nombre, gmv, marca_prod, es_cierre, col_señal, hoy):
    """Devuelve (nivel1, nivel2, nivel3, canal) para una marca. nivel2,
    nivel3 y canal quedan en None si la marca no llega a ese nivel o no
    tiene un contacto real de la palanca del que sacar el canal."""
    contacto_palanca = marca_prod[
        (marca_prod["¿Contactado?"] == "SI") & (marca_prod[col_señal] == "SI")
    ] if not marca_prod.empty else marca_prod

    canal, fecha_inicio = None, None
    if not contacto_palanca.empty:
        cp_ord = contacto_palanca.sort_values("Date")
        fecha_inicio = cp_ord["Date"].min()
        primer_medio = cp_ord.iloc[0].get("Medio de Contacto") if "Medio de Contacto" in cp_ord.columns else None
        canal = _canal_desde_medio(primer_medio)

    if es_cierre:
        return "Contactado", "Pipeline", "Cierre", canal or "Gmail"

    if marca_prod.empty:
        return "Sin Gestionar", None, None, None

    if not contacto_palanca.empty:
        dias = dias_habiles_entre(fecha_inicio, hoy) or 0
        tiene_no_activo = (marca_prod.get("Tipo Never Ads") == "No activo").any() \
            if "Tipo Never Ads" in marca_prod.columns else False
        if tiene_no_activo and dias > UMBRAL_RECHAZO_HABILES:
            return "Contactado", "Rechazado", None, canal
        nivel3 = "Caliente" if dias <= UMBRAL_CALIENTE_HABILES else "Frío"
        return "Contactado", "Pipeline", nivel3, canal

    if (marca_prod["¿Contactado?"] == "NO").any():
        return "No Contactado", None, None, None

    return "Sin Gestionar", None, None, None


def _construir_funnel(universo_keys, nombre_map, gmv_map, cierre_keys, prod_farmer, col_señal, hoy):
    filas = []
    for key in universo_keys:
        marca_prod = prod_farmer[prod_farmer["brand_key"] == key]
        n1, n2, n3, canal = _clasificar_marca(
            key, nombre_map.get(key, key), gmv_map.get(key, 0.0),
            marca_prod, key in cierre_keys, col_señal, hoy,
        )
        filas.append({
            "Brand ID": key, "Brand Name": nombre_map.get(key, key),
            "GMV": gmv_map.get(key, 0.0), "N1": n1, "N2": n2, "N3": n3,
            "Canal": canal,
        })
    df = pd.DataFrame(filas, columns=["Brand ID", "Brand Name", "GMV", "N1", "N2", "N3", "Canal"])
    return df.sort_values("GMV", ascending=False).reset_index(drop=True)


def funnel_niveles(df):
    """Arma los 4 niveles listos para pintar. Cada nivel trae su total,
    los segmentos de su barra interna (nombre, n, %), y el subconjunto de
    marcas que le corresponde con su Status para la tabla lateral.

    BUG REAL ENCONTRADO Y CORREGIDO (pedido explícito de Sabas: "no
    coincide la cuenta del pipeline"). La versión anterior mostraba
    "Pipeline 26" en la barra de Contactados, pero la card de Pipeline de
    abajo mostraba total 12 -- un hueco de 14, EXACTO al número de
    Cierre. Causa: a nivel de dato, una marca cerrada seguía con
    N2="Pipeline" (decisión explícita de una vuelta anterior, para que
    Cierre no apareciera como segmento en ningún lado más que su propia
    card) -- pero el conteo de la barra usaba N2 (que SÍ incluye
    cerrados) mientras el total de la card de Pipeline usaba N3 (que los
    EXCLUYE). Dos filtros distintos para "lo mismo" = el hueco visible.
    Con solo 2 cierres el hueco pasaba casi desapercibido; con 14 se
    volvió imposible de no notar -- por eso apareció recién ahora, no es
    que se haya roto de nuevo.

    FIX: "Pipeline" en la barra de Contactados ahora se cuenta igual que
    en la card de abajo (N3 en Caliente/Frío, sin cerrados), y se agrega
    "Cerrado" como tercer segmento explícito de esa misma barra -- así
    Contactados = Pipeline + Rechazado + Cerrado vuelve a sumar exacto
    (regla de oro), y "Pipeline X" en la barra de arriba es SIEMPRE el
    mismo número que el total de la card de Pipeline debajo, sin
    excepción.
    """
    n_contactado = int((df["N1"] == "Contactado").sum())
    n_cierre = int((df["N3"] == "Cierre").sum())
    n_pipeline = int(df["N3"].isin(["Caliente", "Frío"]).sum())
    n_rechazado = int((df["N2"] == "Rechazado").sum())

    def segs(pares):
        total = sum(n for _, n in pares) or 1
        return [{"label": lab, "n": n, "pct": n / total * 100} for lab, n in pares]

    base_tbl = df.assign(Status=df["N1"])
    cont_tbl = df[df["N1"] == "Contactado"].assign(
        Status=lambda d: d["N3"].where(d["N3"] == "Cierre", d["N2"]).fillna("Pipeline"))
    pipe_tbl = df[df["N3"].isin(["Caliente", "Frío"])].assign(Status=df["N3"])
    cierre_tbl = df[df["N3"] == "Cierre"].assign(Status="Cerrado")
    # Inversión (Pipeline) y Cerrado (Cierre) se resuelven en la UI --
    # ahí es donde se sabe si esta tabla es de Ads o de Markdown y se
    # tienen a mano los rangos de RECOMMENDED BUDGETS y el %CVR.

    return [
        {
            "key": "base", "titulo": "Base prospectada", "total": len(df),
            "sub": "100%",
            "segmentos": segs([(l, int((df["N1"] == l).sum())) for l in FUNNEL_ORDEN_L1]),
            "tabla": base_tbl,
        },
        {
            "key": "contactado", "titulo": "Contactados", "total": n_contactado,
            "sub": f"{(n_contactado / len(df) * 100) if len(df) else 0:.1f}% de la base",
            "segmentos": segs([("Pipeline", n_pipeline), ("Rechazado", n_rechazado), ("Cerrado", n_cierre)]),
            "tabla": cont_tbl,
        },
        {
            "key": "pipeline", "titulo": "Pipeline", "total": n_pipeline,
            "sub": f"de {n_contactado} contactados",
            "segmentos": segs([
                ("Caliente", int((df["N3"] == "Caliente").sum())),
                ("Frío", int((df["N3"] == "Frío").sum())),
            ]),
            "tabla": pipe_tbl,
        },
        {
            "key": "cierre", "titulo": "Cierre", "total": n_cierre,
            "sub": f"{(n_cierre / len(df) * 100) if len(df) else 0:.1f}% de la base",
            "segmentos": [], "tabla": cierre_tbl,
        },
    ]


@st.cache_data(show_spinner=False)
def funnel_ads(ads_df, productivity_df, checkout_df, farmer_email, gmv_map=None, universo_path=None, hoy=None):
    hoy = pd.Timestamp(hoy) if hoy is not None else pd.Timestamp.now().normalize()
    gmv_map = gmv_map or {}
    prod = productivity_df[productivity_df["Farmer"] == farmer_email].copy()
    prod = _asegurar_brand_key(prod)

    d = ads_df.copy()
    if d.empty or "BRAND" not in d.columns:
        return pd.DataFrame(columns=["Brand ID", "Brand Name", "GMV", "N1", "N2", "N3", "Canal"])
    d["brand_key"] = d["BRAND"].apply(_brand_key)
    d["att_pct"] = d["% Att. Bookings"].apply(_parse_pct)
    califican = d[(d["att_pct"] == 0) & (d["brand_key"] != "")]
    nuevos = dict(zip(califican["brand_key"], califican["BRAND"]))
    nombre_map = actualizar_universo_mensual(nuevos, universo_path) if universo_path else nuevos

    cierre_keys = set()
    chk = checkout_df
    if not chk.empty and {"FARMER", "Tipo de Contratacion", "brand_key"}.issubset(chk.columns):
        cierre_keys = set(chk[chk["FARMER"] == farmer_email]["brand_key"])

    # BUG REAL ENCONTRADO (pedido explícito de Sabas: "no me está dando
    # los datos reales" -- Checkout tenía 14 cierres reales, el funnel
    # solo mostraba 2). Causa: el universo se arma SOLO con marcas que
    # en este momento tienen % Att. Bookings == 0. Una marca que cierra
    # empieza a atribuir de verdad -- su % deja de ser 0 casi por
    # definición -- así que sale del filtro crudo justo cuando más
    # importa que se quede. El "universo acumulado" en disco existía
    # para blindar esto, pero es un archivo local efímero (se puede
    # perder en cualquier redeploy de Streamlit Cloud, ya avisado en el
    # README) -- si se pierde, la marca desaparece de la iteración
    # completa y ni siquiera llega a compararse contra Checkout.
    # FIX: cualquier marca que SÍ cerró (aparece en Checkout, la fuente
    # de verdad, que se lee fresca del Excel cada vez, no depende de
    # ningún archivo persistido) entra al universo sin importar si el
    # acumulado la tiene o qué diga su % de hoy. El nombre, si no está
    # en nombre_map, se busca en la propia hoja ADS completa (sin
    # filtrar por %) como respaldo.
    faltantes = cierre_keys - set(nombre_map.keys())
    if faltantes:
        nombre_ads_completo = dict(zip(d["brand_key"], d["BRAND"]))
        nombre_prod = dict(zip(prod["brand_key"], prod["Brand"])) if "Brand" in prod.columns else {}
        for k in faltantes:
            nombre_map[k] = nombre_ads_completo.get(k) or nombre_prod.get(k) or k

    return _construir_funnel(sorted(nombre_map), nombre_map, gmv_map, cierre_keys, prod, "Ads", hoy)


@st.cache_data(show_spinner=False)
def funnel_md(md_df, productivity_df, checkout_df, farmer_email, gmv_map=None, universo_path=None, hoy=None):
    hoy = pd.Timestamp(hoy) if hoy is not None else pd.Timestamp.now().normalize()
    gmv_map = gmv_map or {}
    prod = productivity_df[productivity_df["Farmer"] == farmer_email].copy()
    prod = _asegurar_brand_key(prod)

    d = md_df.copy()
    if d.empty or "BRAND ID" not in d.columns:
        return pd.DataFrame(columns=["Brand ID", "Brand Name", "GMV", "N1", "N2", "N3", "Canal"])
    d["brand_key"] = d["BRAND ID"].apply(_brand_key)
    d["md_pct"] = d["MARKDOWN %"].apply(_parse_pct)
    califican = d[(d["md_pct"].isna() | (d["md_pct"] == 0)) & (d["brand_key"] != "")]
    nuevos = dict(zip(califican["brand_key"], califican["BRAND NAME"]))
    nombre_map = actualizar_universo_mensual(nuevos, universo_path) if universo_path else nuevos

    # Cierre de MD: la aceptación vive en Productivity, no en Checkout.
    cierre_keys = set(prod[prod["¿Se aceptó lo ofrecido?"] == "Sí"]["brand_key"])

    # Mismo blindaje que funnel_ads (ver esa función para el detalle
    # completo del bug): una marca que acepta MD probablemente deja de
    # tener Markdown % en 0/vacío, así que puede caer fuera del filtro
    # crudo justo al cerrar. Se garantiza que entre igual, con el nombre
    # de respaldo desde la propia hoja MD completa.
    faltantes = cierre_keys - set(nombre_map.keys())
    if faltantes:
        nombre_md_completo = dict(zip(d["brand_key"], d["BRAND NAME"]))
        nombre_prod = dict(zip(prod["brand_key"], prod["Brand"])) if "Brand" in prod.columns else {}
        for k in faltantes:
            nombre_map[k] = nombre_md_completo.get(k) or nombre_prod.get(k) or k

    return _construir_funnel(sorted(nombre_map), nombre_map, gmv_map, cierre_keys, prod, "Markdown", hoy)


# ═════════════════════════════════════════════════════════════
# FUNNEL DE CHURN (3 niveles) -- estructura propia, pedida explícitamente
# ═════════════════════════════════════════════════════════════
#   Nivel 1 -- Prospectados = PW1 + Churn (hoja CHURN).
#              Barra interna: Contactado / No Contactado / Sin Gestionar.
#   Nivel 2 -- Contactados = los de arriba con contacto en Productivity.
#              Barra interna: Se reactiva (On Hold=SI) / Cerrado
#              permanente (On Hold=NO).
#   Nivel 3 -- Retenidos = solo los que figuran como "se reactiva".

@st.cache_data(show_spinner=False)
def funnel_churn(churn_df, productivity_df, farmer_email, gmv_map=None):
    gmv_map = gmv_map or {}
    ch = churn_df.copy()
    if ch.empty or "COUNTRY_BRAND_ID" not in ch.columns:
        return pd.DataFrame(columns=["Brand ID", "Brand Name", "GMV", "N1", "N2", "Estado Churn"])
    if "FARMER" in ch.columns:
        local = str(farmer_email).split("@")[0].strip().lower()
        ch = ch[ch["FARMER"].astype(str).str.strip().str.lower() == local]
    ch["brand_key"] = ch["COUNTRY_BRAND_ID"].apply(_brand_key)

    # BUG ENCONTRADO Y CORREGIDO: antes esto filtraba prod por
    # Churn=="SI", pero las marcas que hoy están en la hoja CHURN casi no
    # tienen filas marcadas así (esas filas corresponden a gestiones de
    # churn de OTRAS marcas, en otro momento). La spec dice "cuántos de
    # los PW1 y los Churn tienen contacto en productivity" -- sin
    # restringir el tipo de fila -- así que se mira TODA la actividad de
    # la marca. Con el filtro viejo, el nivel 2 daba 0 siempre.
    prod = productivity_df[productivity_df["Farmer"] == farmer_email].copy()
    prod = _asegurar_brand_key(prod)

    filas = []
    for _, row in ch.iterrows():
        key = row["brand_key"]
        mp = prod[prod["brand_key"] == key]
        estado_churn = "PW1" if row.get("Estado Actual") == "Prevention W1" else "Churn"

        if mp.empty:
            n1, n2 = "Sin Gestionar", None
        elif (mp["¿Contactado?"] == "SI").any():
            n1 = "Contactado"
            n2 = "Se reactiva" if (mp["On Hold"] == "SI").any() else "Cerrado permanente"
        elif (mp["¿Contactado?"] == "NO").any():
            n1, n2 = "No Contactado", None
        else:
            n1, n2 = "Sin Gestionar", None

        filas.append({
            "Brand ID": key, "Brand Name": row.get("BRAND_NAME", key),
            "GMV": gmv_map.get(key, 0.0), "N1": n1, "N2": n2,
            "Estado Churn": estado_churn,
        })

    df = pd.DataFrame(filas, columns=["Brand ID", "Brand Name", "GMV", "N1", "N2", "Estado Churn"])
    return df.sort_values("GMV", ascending=False).reset_index(drop=True)


def funnel_churn_niveles(df):
    n_cont = int((df["N1"] == "Contactado").sum())
    n_ret = int((df["N2"] == "Se reactiva").sum())

    def segs(pares):
        total = sum(n for _, n in pares) or 1
        return [{"label": lab, "n": n, "pct": n / total * 100} for lab, n in pares]

    return [
        {
            "key": "prospectados", "titulo": "Prospectados (PW1 + Churn)", "total": len(df),
            "sub": "100%",
            "segmentos": segs([(l, int((df["N1"] == l).sum())) for l in FUNNEL_ORDEN_L1]),
            "tabla": df.assign(Status=df["N1"]),
        },
        {
            "key": "contactados", "titulo": "Contactados", "total": n_cont,
            "sub": f"{(n_cont / len(df) * 100) if len(df) else 0:.1f}% de prospectados",
            "segmentos": segs([
                ("Se reactiva", int((df["N2"] == "Se reactiva").sum())),
                ("Cerrado permanente", int((df["N2"] == "Cerrado permanente").sum())),
            ]),
            "tabla": df[df["N1"] == "Contactado"].assign(Status=lambda d: d["N2"]),
        },
        {
            "key": "retenidos", "titulo": "Retenidos", "total": n_ret,
            "sub": f"{(n_ret / len(df) * 100) if len(df) else 0:.1f}% de prospectados",
            "segmentos": [],
            "tabla": df[df["N2"] == "Se reactiva"].assign(Status="Se reactiva"),
        },
    ]


# ═════════════════════════════════════════════════════════════
# INVERSIÓN RECOMENDADA (Pipeline) y % CERRADO (Cierre)
# ═════════════════════════════════════════════════════════════
# Pedido explícito de Sabas: en la tabla de Pipeline, columna "Inversión"
# con el % recomendado según la hoja RECOMMENDED BUDGETS -- para Ads,
# por el GMV convertido a ARS (la tabla lo muestra en USD); para
# Markdown, por el %CVR de la hoja %CVR. En la tabla de Cierre, columna
# "Cerrado" con el % cerrado según la columna Presupuesto de Checkout,
# venga como % o como $.

def _parse_rango(texto):
    """'<145000' / '145000 - 725000' / '> 7250000' / '<10%' / '10% - 15 %'
    -> (lo, lo_incl, hi, hi_incl). None en lo/hi = sin piso/techo."""
    s = str(texto).strip()
    if s.startswith("<"):
        hi = float(re.sub(r"[^\d.,]", "", s).replace(",", "."))
        return None, False, hi, False
    if s.startswith(">"):
        lo = float(re.sub(r"[^\d.,]", "", s).replace(",", "."))
        return lo, False, None, False
    if "-" in s:
        izq, der = s.split("-", 1)
        lo = float(re.sub(r"[^\d.,]", "", izq).replace(",", "."))
        hi = float(re.sub(r"[^\d.,]", "", der).replace(",", "."))
        return lo, True, hi, True
    return None, False, None, False


def _pct_por_rango(valor, rangos):
    """`rangos`: lista de (lo, lo_incl, hi, hi_incl, pct). Devuelve el
    pct del primer rango que contiene `valor`, o None si no calza en
    ninguno (valor NaN/None, o tabla de rangos vacía)."""
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None
    for lo, lo_incl, hi, hi_incl, pct in rangos:
        ok_lo = lo is None or (valor >= lo if lo_incl else valor > lo)
        ok_hi = hi is None or (valor <= hi if hi_incl else valor < hi)
        if ok_lo and ok_hi:
            return pct
    return None


@st.cache_data(show_spinner=False)
def load_recommended_budgets(file_like_or_path):
    """Lee la hoja 'RECOMMENDED BUDGETS' -- dos tablas apiladas en las
    mismas 2 columnas: rangos de ARS$ VENTAS -> % recomendado (Ads), y
    rangos de %CVR -> % recomendado (Markdown), separadas por una fila
    con el encabezado real de la segunda tabla ('%CVR'). Devuelve
    (rangos_ads, rangos_md), cada uno listo para _pct_por_rango. Si la
    hoja no existe o no calza el formato esperado, devuelve ([], [])
    -- nunca revienta la app."""
    try:
        xls = pd.ExcelFile(file_like_or_path) if not isinstance(file_like_or_path, pd.ExcelFile) else file_like_or_path
        lower_map = {name.strip().lower(): name for name in xls.sheet_names}
        real_name = lower_map.get("recommended budgets")
        if not real_name:
            return [], []
        raw = pd.read_excel(xls, sheet_name=real_name, header=None)
    except (OSError, ValueError, KeyError):
        return [], []

    split_idx = None
    for i, val in enumerate(raw[0]):
        if isinstance(val, str) and val.strip() == "%CVR":
            split_idx = i
            break

    rangos_ads, rangos_md = [], []
    tabla1 = raw.iloc[1:split_idx] if split_idx is not None else raw.iloc[1:]
    for _, row in tabla1.iterrows():
        rango, pct = row[0], row[1]
        if pd.isna(rango) or pd.isna(pct):
            continue
        lo, lo_incl, hi, hi_incl = _parse_rango(rango)
        rangos_ads.append((lo, lo_incl, hi, hi_incl, float(pct) * 100))

    if split_idx is not None:
        tabla2 = raw.iloc[split_idx + 1:]
        for _, row in tabla2.iterrows():
            rango, pct = row[0], row[1]
            if pd.isna(rango) or pd.isna(pct):
                continue
            lo, lo_incl, hi, hi_incl = _parse_rango(rango)
            rangos_md.append((lo, lo_incl, hi, hi_incl, float(pct) * 100))

    return rangos_ads, rangos_md


def inversion_ads_pct(gmv_usd, rangos_ads, tasa_ars=TASA_USD_ARS):
    """% de inversión recomendado para Ads: el GMV (USD, como lo muestra
    la tabla) se convierte a ARS -- moneda de los rangos de la hoja --
    y se ubica en el rango que corresponda."""
    if gmv_usd is None:
        return None
    return _pct_por_rango(float(gmv_usd) * tasa_ars, rangos_ads)


def inversion_md_pct(cvr_pct, rangos_md):
    """% de inversión recomendado para Markdown, según el %CVR de la
    marca (ya expresado como porcentaje, ej. 17.81)."""
    return _pct_por_rango(cvr_pct, rangos_md)


def nombre_md_lookup(md_df):
    """{brand_key: nombre crudo (sin ID)} desde MD.BRAND NAME -- MD es la
    única hoja del cruce cuyo nombre de marca coincide, tal cual, con
    los nombres que trae la hoja %CVR (que no incluye Brand ID)."""
    d = md_df.copy()
    if d.empty or "BRAND ID" not in d.columns or "BRAND NAME" not in d.columns:
        return {}
    d["brand_key"] = d["BRAND ID"].apply(_brand_key)
    return dict(zip(d["brand_key"], d["BRAND NAME"]))


def _norm_nombre(s):
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def cvr_lookup(cvr_df):
    """Lee la hoja '%CVR' -- export crudo cuyo encabezado real
    (Métrica | Brand Name | Valor | vs LM) vive en la primera fila de
    datos, no en el header de pandas. Devuelve {nombre_normalizado:
    %CVR} (ej. 0.1781 -> 17.81)."""
    d = cvr_df
    if d is None or d.empty or d.shape[1] < 3:
        return {}
    cols = list(d.columns)
    nombre_col, valor_col = cols[1], cols[2]
    out = {}
    for _, row in d.iloc[1:].iterrows():
        nombre, valor = row.get(nombre_col), row.get(valor_col)
        if pd.isna(nombre) or pd.isna(valor):
            continue
        try:
            out[_norm_nombre(nombre)] = float(valor) * 100
        except (TypeError, ValueError):
            continue
    return out


@st.cache_data(show_spinner=False)
def load_cvr(file_like_or_path):
    """Lee la hoja '%CVR' del cruce por nombre (case-insensitive). Hoja
    ausente -> DataFrame vacío, no revienta."""
    try:
        xls = pd.ExcelFile(file_like_or_path) if not isinstance(file_like_or_path, pd.ExcelFile) else file_like_or_path
        lower_map = {name.strip().lower(): name for name in xls.sheet_names}
        real_name = lower_map.get("%cvr")
        if not real_name:
            return pd.DataFrame()
        return pd.read_excel(xls, sheet_name=real_name)
    except (OSError, ValueError, KeyError):
        return pd.DataFrame()


@st.cache_data(show_spinner=False)
def cvr_por_brand_key(md_df, cvr_df):
    """{brand_key: %CVR} combinando el nombre crudo de MD con el valor de
    la hoja %CVR -- puente entre las dos hojas, que no comparten Brand ID."""
    nombre_map = nombre_md_lookup(md_df)
    cvr_map = cvr_lookup(cvr_df)
    return {key: cvr_map.get(_norm_nombre(nombre)) for key, nombre in nombre_map.items()}


def _parse_presupuesto_valor(valor_crudo, es_pct=False):
    """
    Devuelve (tipo, valor) del 'Presupuesto' de Checkout TAL CUAL viene
    -- pedido explícito de Sabas: "colocas el valor independiente que
    sea $ o %, no lo conviertas todo a porcentaje" (convertir a % del
    GMV daba números sin sentido en marcas de GMV chico -- ej. $40.000
    de presupuesto sobre $10 de GMV daba "275.9%").

    BUG REAL ENCONTRADO (pedido explícito de Sabas: "el cierre me está
    mostrando $0, ¿es real?"): antes esto decidía "pct" vs "monto"
    buscando el símbolo "%" en el texto del valor. Pero pandas nunca
    trae ese símbolo -- vive en el FORMATO VISUAL de la celda de Excel
    ('0%'), no en el valor numérico crudo que devuelve read_excel(). Una
    celda con 10% real y una con $0.10 de pesos llegan AMBAS como el
    float 0.1, indistinguibles con el método viejo -- por eso $0.1
    pesos ÷ 1450 (tasa USD) redondeaba a $0 para cada fila que en
    realidad era un %.

    `es_pct` ahora viene del formato REAL de la celda (ver
    _detectar_formato_pct en load_cruce, leído con openpyxl) -- ya no se
    adivina con texto.
    """
    if pd.isna(valor_crudo):
        return None
    s = str(valor_crudo).strip()
    if not s:
        return None
    if es_pct or "%" in s:
        valor = float(valor_crudo) * 100 if not isinstance(valor_crudo, str) else _parse_pct(s)
        return ("pct", valor)
    return ("monto", _parse_money(s))


@st.cache_data(show_spinner=False)
def presupuesto_valor_por_brand(checkout_df, farmer_email):
    """{brand_key: (tipo, valor)} desde Checkout.Presupuesto para este
    Account Manager. Con más de una fila de Checkout por marca, se
    queda con la más reciente (ordenado por Fecha antes de recorrer)."""
    chk = checkout_df
    if chk is None or chk.empty or not {"FARMER", "brand_key", "Presupuesto"}.issubset(chk.columns):
        return {}
    sub = chk[chk["FARMER"] == farmer_email]
    if "Fecha" in sub.columns:
        sub = sub.sort_values("Fecha")
    tiene_flag = "_presupuesto_es_pct" in sub.columns
    out = {}
    for _, row in sub.iterrows():
        es_pct = bool(row.get("_presupuesto_es_pct", False)) if tiene_flag else False
        val = _parse_presupuesto_valor(row.get("Presupuesto"), es_pct=es_pct)
        if val is not None:
            out[row["brand_key"]] = val
    return out
